
# распределённая сила
# круглое сечение
import math

from matplotlib import pyplot as plt
from openpyxl import Workbook

E = 130e9

F = 0.6e-6    # ньютоны
L = 600e-6    # метры
W = 4e-6      # метры
H = 4e-6      # метры

r = W / 2
q = F / L


def get_phi(x):
    return (2*q)*(L**3 - x**3)/(3*E*math.pi*(r**4))


def get_z(x):
    return (-4*q) * ((x**4) - 4*(L**3)*x + 3*(L**4)) / (24*E*math.pi*(r**4))


print(f'q = {q} Н/м')
print(f'phi_max = {get_phi(0)}')
print(f'z_max = {get_z(0)}')

wb = Workbook()



ws = wb.create_sheet('φ')

x1, y1 = [], []
i = 0
while i <= 2*L:
    x1.append(i)
    y1.append(get_phi(i))
    i += L/50

count = 1
for i, j in zip(x1, y1):
    ws[f'A{count}'] = i
    ws[f'B{count}'] = j
    count += 1




ws = wb.create_sheet('z')

x2, y2 = [], []
i = 0
while i <= 2*L:
    x2.append(i)
    y2.append(get_z(i))
    i += L/50


count = 1
for i, j in zip(x2, y2):
    ws[f'A{count}'] = i
    ws[f'B{count}'] = j
    count += 1



wb.save('rgr4.xlsx')

plt.plot(x1, y1)
plt.scatter(x1, y1, s=5)
plt.xlabel('x, м')
plt.ylabel('φ')
plt.grid()

plt.figure()
plt.plot(x2, y2)
plt.scatter(x2, y2, s=5)
plt.xlabel('x, м')
plt.ylabel('z, м')
plt.grid()

plt.show()
