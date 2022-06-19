from openpyxl import Workbook


E = 270e9

L = 600e-6
W = 4e-6
H = 0.7e-6
F = 0.6e-6


def get_y(form, ampl):
    res = 12/E
    if form != 'F':
        res *= F
    if form != 'l':
        res *= L*L*L
    if form != 'w':
        res /= W
    if form != 'h':
        res /= H*H*H

    return res * get(form, ampl)


def get(form, amp):
    if form == 'F':
        return F * amp
    if form == 'l':
        return (L * amp) * (L * amp) * (L * amp)
    if form == 'h':
        return 1 / (H * amp) / (H * amp) / (H * amp)
    if form == 'w':
        return 1 / (W * amp)


def make_arr(form):
    res = [[], []]
    for i in range(90, 110+2, 2):
        res[0].append(i/100)
        res[1].append(get_y(form, i/100))
    return res


modes = ['F', 'l', 'h', 'w']

wb = Workbook()

for mode in modes:
    ws = wb.create_sheet(mode)
    arr = make_arr(mode)
    for j in range(len(arr[0])):
        ws['A'+str(j+1)] = arr[0][j]
        ws['B'+str(j+1)] = arr[1][j]

wb.save('rgr2.xlsx')

