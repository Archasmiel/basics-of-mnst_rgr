from matplotlib import pyplot as plt
from math import *

from openpyxl import Workbook

alphaT = 5.1e-6         # 1/°C
delta_l1_max = 6e-10    # m
temp = [70, 120]        # °C

delta_T_max = temp[1] - temp[0]
l2 = delta_l1_max/delta_T_max/alphaT


def get_y(dtemp: float):
    return l2*sin(sqrt(6*alphaT*dtemp))


print(f'l2 = {l2}')

wb = Workbook()
ws = wb.active

x, y = [temp[0]], [0]
for i in range(temp[0]+1, temp[1]+1, 1):
    x.append(i)
    y.append(get_y(i-temp[0]))

count = 1
for i, j in zip(x, y):
    ws[f'A{count}'] = i
    ws[f'B{count}'] = j
    count += 1

wb.save('rgr3.xlsx')

plt.plot(x, y)
plt.scatter(x, y, s=5)
plt.xlabel('T, °C')
plt.ylabel('y, м')
plt.grid()
plt.show()
