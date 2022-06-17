from matplotlib import pyplot as plt
from math import *

from openpyxl import Workbook

alphaT = 5.1
delta_l1_max = 6e-4
temp = [70, 80, 90, 100, 110, 120]

delta_T_max = temp[-1]
l2 = delta_l1_max/delta_T_max/alphaT

epsi_deform_max = alphaT*delta_T_max


def get_angle(dtemp: float):
    return sqrt(24 * alphaT * dtemp)


def get_radius(dtemp: float):
    return l2 * sqrt(alphaT * dtemp) / sqrt(24)


def get_y(dtemp: float):
    return l2*l2/2/get_radius(dtemp)


x, y = [], []
for i in range(temp[0]+1, temp[-1]+1, 2):
    x.append(i)
    y.append(get_y(i))
    

print()
print(get_angle(delta_T_max))
print(epsi_deform_max)

wb = Workbook()
ws = wb.active

count = 1
for i, j in zip(x, y):
    ws[f'A{count}'] = i
    ws[f'B{count}'] = j
    count += 1

wb.save('rgr3.xlsx')

plt.plot([min(x), max(x)], [min(y), min(y)])
plt.plot([min(x), min(x)], [min(y), max(y)])
plt.scatter(x, y, s=1)
plt.grid()
plt.show()
