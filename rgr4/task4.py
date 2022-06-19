from math import pi
from info import task4
from matplotlib import pyplot as plt
from openpyxl import Workbook

var = open('../variant.txt', 'r').readline().strip()
inf1, inf2, material, L, W, H, F = task4.get(var)

E = material.get('E')


def get_j():
    if inf1 == 'kr':
        return pi*((W/2)**4)/4
    return (H**3)*W/12


def get_phi(x):
    J = get_j()
    if inf2 == 'zos':
        return F * (L ** 2 - x ** 2) / (2 * E * J)
    if inf2 == 'roz':
        q = F / L
        return q * (L ** 3 - x ** 3) / (6 * E * J)


def get_z(x):
    J = get_j()
    if inf2 == 'roz':
        q = F / L
        return -q * ((x ** 4) - 4 * (L ** 3) * x + 3 * (L ** 4)) / (24 * E * J)
    if inf2 == 'zos':
        return -F * ((x**3)/6 - (L**2)*x/2 + (L**3)/3) / (E * J)


def print_nums():
    wb = Workbook()
    del wb['Sheet']

    ws = wb.create_sheet('φ')

    x1, y1 = [], []
    i = 0
    while i <= 2 * L:
        x1.append(i)
        y1.append(get_phi(i))
        i += L / 50

    count = 1
    for i, j in zip(x1, y1):
        ws[f'A{count}'] = i
        ws[f'B{count}'] = j
        count += 1

    ws = wb.create_sheet('z')

    x2, y2 = [], []
    i = 0
    while i <= 2 * L:
        x2.append(i)
        y2.append(get_z(i))
        i += L / 50

    count = 1
    for i, j in zip(x2, y2):
        ws[f'A{count}'] = i
        ws[f'B{count}'] = j
        count += 1

    wb.save('rgr4.xlsx')

    plt.plot(x1, y1)
    plt.scatter(x1, y1, s=5)
    plt.xlabel('x, м')
    plt.ylabel('φ')
    plt.grid()

    plt.figure()
    plt.plot(x2, y2)
    plt.scatter(x2, y2, s=5)
    plt.xlabel('x, м')
    plt.ylabel('z, м')
    plt.grid()

    plt.show()


def print_info():
    print(f'phi_max = {get_phi(0)}')
    print(f'z_max = {get_z(0)}')


print_info()
print_nums()
